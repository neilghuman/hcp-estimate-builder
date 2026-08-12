import openpyxl
from pathlib import Path
p = Path(r"C:\Users\neil\Downloads\a0ba9d11-944c-48c7-82fa-a69e41b1cd30.xlsx")
print(f"exists={p.exists()} size={p.stat().st_size if p.exists() else 0}")
wb = openpyxl.load_workbook(p, data_only=True)
print('sheets:', ', '.join(wb.sheetnames))
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n=== {name} ===")
    print(f"rows={ws.max_row} cols={ws.max_column}")
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 6), values_only=True):
        print(row)
